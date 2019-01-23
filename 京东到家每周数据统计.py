# @Date    : 2018-06-11 20:07:43
# @Author  : Hume (102734075@qq.com)
# @Link    : https://humecry.wordpress.com/
# @Version : 1.1
# @Description：获取京东到家与美团外卖各门店上周数据并添加进Excel

import requests
import time
from openpyxl.styles import Font, colors, Alignment, Border, Side, numbers, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import datetime
from datetime import timedelta
import pandas as pd

# 当前时间
now = datetime.datetime.now()
# 上周的第一天周日，最后一天周六
last_week_start = now - timedelta(days=7+now.isoweekday())
last_week_end = now - timedelta(days=1+now.isoweekday())
'''
获取京东到家上周数据
'''
def getJD(headers, shops):
	dfs = []
	# 添加上周京东到家数据
	for shop in shops.values():
		timeDate = last_week_start.strftime("%Y-%m-%d") + "~" + last_week_end.strftime("%Y-%m-%d")
		yearWeek = time.strftime("%Y{Y}%W{W}", time.localtime(time.time()-7*24*60*60)).format(Y="年", W="周")
		params = {
		    "venderId": "320695",
		    "timeRangeType": "2",
		    "weekId": yearWeek,
		    "dateWeek": yearWeek,
		    "shopIdListStr": shop
		}
		# 获取主要数据
		url = 'https://dc-store.jd.com/operation/queryShopOperationData'
		response = requests.get(url, headers=headers, params=params, timeout=10).json()
		data = response['operationDataResponseDTOs'][0]
		# 获取环比
		url2 = 'https://dc-store.jd.com/operation/queryData'
		response2 = requests.get(url2, headers=headers, params=params, timeout=10).json()
		row = [timeDate, data['shopName'], data['browseCnt'], data['totalVisitCnt'], data['validOrderCnt'], data['takeRate']/100, data['orderTotalAmtz'], data['perTicketSales'], 0 if response2['validOrderCountRelativeRatio']=='--' else response2['validOrderCountRelativeRatioSign']*float(response2['validOrderCountRelativeRatio'])/100]
		dfs.append(pd.DataFrame([row], columns=['时间','门店', '浏览量', '访客数', '有效订单数', '转化率', 'GMV成交额', '客单价', '环比']))
	df1 = dfs.pop()
	for value in dfs:
		df1 = pd.concat([df1, value], ignore_index=True)
	return df1
'''
获取美团外卖华森店上周数据
'''
def getMeiTuan(headers):
	timeDate = last_week_start.strftime("%Y-%m-%d") + "~" + last_week_end.strftime("%Y-%m-%d")
	params = {
	    'wmPoiId': '3578168',
	    'beginTime': last_week_start.strftime("%Y%m%d"),
	    'endTime': last_week_end.strftime("%Y%m%d"),
	}
	# 获取主要数据
	url = 'http://waimaieapp.meituan.com/bizdata/businessStatisticsV3/single/hisOverview'
	response = requests.get(url, headers=headers, params=params, timeout=10).json()
	data = response['data']
	params2 = {
		'wmPoiId': '3578168',
		'recentDays': '7'
	}
	# # 获取环比
	url2 = 'http://waimaieapp.meituan.com/bizdata/flowanalysisV2/flow/overview'
	response2 = requests.get(url2, headers=headers, params=params2, timeout=10).json()
	data2 = response2['data']['flowOverviewInfo']
	row = [timeDate, '美团外卖-华森店', data2['exposureNum'], data2['visitNum'], data['effectiveOrders'], data['effectiveOrders']/data2['visitNum'], data['turnover'], data['turnover']/data['effectiveOrders'], (data['effectiveOrders']-data['effectiveOrdersLastPeriod'])/data['effectiveOrdersLastPeriod']]
	df = (pd.DataFrame([row], columns=['时间','门店', '浏览量', '访客数', '有效订单数', '转化率', 'GMV成交额', '客单价', '环比']))
	return df

# 要修改的Excel文件
ExcelFilename = '京东到家每周数据.xlsx' 
# 获取Excel数据
wb = load_workbook(ExcelFilename)
sheet = wb.active
data = sheet.values
cols = next(data)[:] # 获取表头
ExcelDf = pd.DataFrame(data, columns=cols)
# 京东到家
JDheaders = {
	'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36',
	'Cookie': 'store.o2o.jd.com1=Q27JKYFDKBBBHSVYSJF2I4UZXJJ7EZTZHJO2A5ZZFEBIUBB3EULCOLDJVEZEP5D7RL4LD27XZJPLYGU4HL4KYQC3LGJHW2QILGDB7IFG52AKHH76WR3WSAD4GIAQ54THKL6XJFIJ63DBAHTSJN4KV4SRA77F36WDQ53Q2GNX6JEUQUYXO36ILVX7Z4M63HXEBZIXN2GSKCSCWIYLCCX4NFQ35JMRQWGUV2ABVVADUHDL242OR3DMHY5S6RN33VRHSHZBQN5H3C4TCLUKVQGVYJFI7A',
}
# 京东到家的商铺与编号
JDshops = {'塔埔店': '11728789', '绿苑店': '11728788'}
JDdf = getJD(JDheaders, JDshops)
# 美团外卖华森店
MTheaders = {
	'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36',
	'Cookie': 'token=0rf-tqbVG8RwWh1tRSYGcqbkWXBXqYWMwUvlS3cX1ty4*; acctId=31330952',
}
MTdf = getMeiTuan(MTheaders)
# 拼接数据
df = pd.concat([JDdf, MTdf], ignore_index=True)
df = pd.concat([ExcelDf, df], ignore_index=True)
df.drop_duplicates(['时间', '门店'],inplace=True)
# 使用pandas进行排序
df.sort_values(['门店', '时间'], inplace=True)
# 删除原Excel表格
wb.remove(sheet)
# # 创建新表
sheet = wb.create_sheet("京东到家周数据")
for r in dataframe_to_rows(df, index=False):
    sheet.append(r)
# 冻结窗格
sheet.freeze_panes = 'A2'
# 设置单元格列宽
sheet.column_dimensions['A'].width = 33
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 15
sheet.column_dimensions['H'].width = 15
sheet.column_dimensions['I'].width = 15
# 单元格边框
border = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'),diagonal=Side(style='thin',color='FF000000'),diagonal_direction=0,outline=Side(style='thin',color='FF000000'),vertical=Side(style='thin',color='FF000000'),horizontal=Side(style='thin',color='FF000000'))
# 所有单元格设置字体
for i in range(sheet.max_row):
	for k in range(sheet.max_column):
		cell = sheet.cell(row=i+1, column=k+1)
		if k==0 or i==0:
			# 设置首行首列加粗
			cell.font = Font(name="Arial", bold=True, size=13)
			# 垂直居中和水平居中
			cell.alignment = Alignment(horizontal='center', vertical='center')
		else:
			# 设置百分比数据格式
			if k==5 or k==8:
				cell.number_format = numbers.FORMAT_PERCENTAGE_00
			cell.font = Font(name="微软雅黑", size=13)
		# 设置单元格边框
		cell.border = border

# 条件格式
redFill = PatternFill(start_color='FF0040', end_color='FF0040', fill_type='solid')
greenFill = PatternFill(start_color='01DF3A', end_color='01DF3A', fill_type='solid')
sheet.conditional_formatting.add('I2:I'+str(len(df)+1), CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=redFill))
sheet.conditional_formatting.add('I2:I'+str(len(df)+1), CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, fill=greenFill))
# 保存Excel
wb.save(filename=ExcelFilename)